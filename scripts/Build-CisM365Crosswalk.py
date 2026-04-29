"""Build CIS M365 crosswalk and SCuBA NIST mapping files for CheckID.

Reads from SecFrame/CIS/:
  - CIS_Microsoft_365_Foundations_Benchmark_v6.0.1.xlsx  (CIS M365 -> safeguards)
  - CIS_Controls_v8.1_Mapping_to_NIST_SP_800-53_Rev5.csv (safeguards -> NIST)
  - scuba-to-nist-sp-800-53-r5-fedramp-high.csv          (SCuBA policy -> NIST)

Produces:
  - data/cis-m365-crosswalk.json    CIS M365 v6 -> NIST via CIS safeguard path
  - data/scuba-nist-mapping.json    SCuBA policy -> NIST (CISA-authoritative)
  - data/framework-titles.json      updated with cis-m365-v6 section

Usage:
    python scripts/Build-CisM365Crosswalk.py
    python scripts/Build-CisM365Crosswalk.py --cis-dir path/to/SecFrame/CIS
"""

import argparse
import csv
import json
import re
from collections import OrderedDict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("ERROR: openpyxl is required. Run: pip install openpyxl")

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_CIS_DIR = REPO_ROOT.parent / "SecFrame" / "CIS"

XLSX_NAME   = "CIS_Microsoft_365_Foundations_Benchmark_v6.0.1.xlsx"
SG_CSV_NAME = "CIS_Controls_v8.1_Mapping_to_NIST_SP_800-53_Rev5.csv"
SCUBA_CSV_NAME = "scuba-to-nist-sp-800-53-r5-fedramp-high.csv"

CROSSWALK_OUT     = REPO_ROOT / "data" / "cis-m365-crosswalk.json"
SCUBA_MAPPING_OUT = REPO_ROOT / "data" / "scuba-nist-mapping.json"
TITLES_PATH       = REPO_ROOT / "data" / "framework-titles.json"

SCHEMA_VERSION = "1.2.0"
# Phase 1 of #347: capture factual + crosswalk metadata from the CIS XLSX
# (Section #, Assessment Status, IG eligibility, CIS Controls + Safeguards by version,
# Default Value, References). CIS-authored prose (Description, Rationale Statement,
# Impact Statement, Remediation Procedure, Audit Procedure, Additional Information)
# is deferred pending licensing resolution — see #347 and docs/SCHEMA-MIGRATION.


# ---------------------------------------------------------------------------
# NIST ID normalisation
# ---------------------------------------------------------------------------

def parse_nist_ids(raw: str, sep: str = ";") -> list[str]:
    """Split and normalise NIST control IDs from a delimited string.

    Handles both semicolon (CIS Controls CSV) and comma (SCuBA CSV) separators.
    Normalises:
      - whitespace before parens: "AC-6 (5)" -> "AC-6(5)"
      - letter suffix on base control: "IA-5c", "IA-5g" -> "IA-5"
      - letter-paren suffix on enhancement: "SC-7(10)(a)" -> "SC-7(10)"
    Deduplicates while preserving first-seen order.
    """
    ids: list[str] = []
    seen: set[str] = set()
    for part in re.split(r"[;,]", raw):
        part = part.strip()
        if not part:
            continue
        part = re.sub(r'\s+\(', '(', part)
        part = re.sub(r'\([a-z]\)$', '', part)          # SC-7(10)(a) -> SC-7(10)
        part = re.sub(r'^([A-Z]+-\d+)[a-z]+$', r'\1', part)  # IA-5c, IA-5g -> IA-5
        if part not in seen:
            seen.add(part)
            ids.append(part)
    return ids


# ---------------------------------------------------------------------------
# CIS safeguard -> NIST lookup
# ---------------------------------------------------------------------------

def load_safeguard_nist(csv_path: Path) -> dict[str, list[str]]:
    """{safeguard_id: [nist_control_ids]} from CIS Controls v8.1 -> NIST CSV."""
    mapping: dict[str, list[str]] = {}
    with open(csv_path, encoding="latin-1", newline="") as fh:
        for row in csv.DictReader(fh):
            sg = row.get("CIS Sub-Control", "").strip()
            nist = row.get("Control Identifier", "").strip()
            if sg and nist:
                mapping.setdefault(sg, [])
                if nist not in mapping[sg]:
                    mapping[sg].append(nist)
    return mapping


# ---------------------------------------------------------------------------
# CIS M365 crosswalk (CIS XLSX + safeguard NIST CSV)
# ---------------------------------------------------------------------------

def _opt_col(headers: tuple, name: str) -> int | None:
    """Find header column by exact name; return None if absent (defensive)."""
    for i, h in enumerate(headers):
        if h == name:
            return i
    return None


def _cell_text(value) -> str:
    """Normalize cell value to trimmed string; '' for None."""
    if value is None:
        return ""
    return str(value).strip()


def _parse_safeguards_cell(value) -> list[str]:
    """Split a CIS Safeguards cell (typically newline- or comma-separated dotted IDs).

    Filters empty / 'None' / non-id tokens. Preserves first-seen order.
    """
    text = _cell_text(value)
    if not text or text.lower() in ("none", "n/a"):
        return []
    parts = re.split(r'[\n,;]+', text)
    out: list[str] = []
    seen: set[str] = set()
    for p in parts:
        t = p.strip()
        if not t or t.lower() in ("none", "n/a"):
            continue
        # Match dotted-numeric safeguard IDs like 6.6, 6.6.1, 4.5
        if not re.match(r'^\d+(\.\d+)*$', t):
            continue
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out


def _parse_ig_flag(value) -> bool:
    """An IG eligibility flag is typically 'X', 'Yes', '✓', or empty.

    Returns True when the cell is non-empty and not a falsy marker.
    """
    text = _cell_text(value).lower()
    if not text:
        return False
    return text not in ("no", "n", "0", "false", "-")


def _parse_references(value) -> list[dict]:
    """Parse References cell into [{url, title}] entries.

    The cell is typically newline-separated URL lines, optionally with title prefix.
    Best-effort — preserves bare URLs as {url: ...} when no title is parseable.
    """
    text = _cell_text(value)
    if not text:
        return []
    refs: list[dict] = []
    for line in re.split(r'[\r\n]+', text):
        line = line.strip()
        if not line:
            continue
        # Match a URL anywhere in the line; if there's prose before/after, treat
        # the prose as the title.
        m = re.search(r'(https?://\S+)', line)
        if not m:
            continue
        url = m.group(1).rstrip('.,;:)')
        title = line.replace(url, "").strip(' -|—:')
        entry: dict = {"url": url}
        if title:
            entry["title"] = title
        refs.append(entry)
    return refs


def _parse_cis_controls_cell(value) -> list[str]:
    """Parse the 'CIS Controls' column — list of v8 control numbers (e.g., '6.6')."""
    text = _cell_text(value)
    if not text or text.lower() in ("none", "n/a"):
        return []
    parts = re.split(r'[\n,;]+', text)
    out: list[str] = []
    seen: set[str] = set()
    for p in parts:
        t = p.strip()
        if not t:
            continue
        # Accept dotted-numeric forms; reject everything else
        if not re.match(r'^\d+(\.\d+)*$', t):
            continue
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out


def build_crosswalk(xlsx_path: Path, sg_nist: dict[str, list[str]]) -> tuple[dict, dict]:
    """Parse CIS M365 XLSX and return (controls_dict, titles_dict).

    Schema v1.2.0 controls_dict shape per recommendation #:
      {
        title, section, sectionNumber, level, license,
        assessmentStatus,                  # "Manual" | "Automated"
        safeguards,                        # legacy: union of v8 safeguards
        nist800_53,
        cisControls,                       # CIS Controls v8 control numbers
        cisSafeguardsByVersion: {
          v8: { ig1: [...], ig2: [...], ig3: [...], applicableIGs: ["IG1"|"IG2"|"IG3"] },
          v7: { ig1: [...], ig2: [...], ig3: [...], applicableIGs: [...] }
        },
        defaultValue,                      # Microsoft factory state — factual
        references                         # [{url, title}] — citation-only, factual
      }

    Phase 1 of #347 deliberately omits CIS-authored prose columns (Description,
    Rationale Statement, Impact Statement, Remediation Procedure, Audit Procedure,
    Additional Information) pending licensing resolution.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Combined Profiles"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]

    def col(name: str) -> int:
        return next(i for i, h in enumerate(headers) if h == name)

    rec_col   = col("Recommendation #")
    title_col = col("Title")
    sg_cols   = [i for i, h in enumerate(headers) if h and "Safeguard" in str(h) and "v8" in str(h)]

    # Optional columns (#347 phase 1 — factual subset only)
    section_num_col   = _opt_col(headers, "Section #")
    assess_col        = _opt_col(headers, "Assessment Status")
    cis_controls_col  = _opt_col(headers, "CIS Controls")
    sg_v8_ig1_col     = _opt_col(headers, "CIS Safeguards 1 (v8)")
    sg_v8_ig2_col     = _opt_col(headers, "CIS Safeguards 2 (v8)")
    sg_v8_ig3_col     = _opt_col(headers, "CIS Safeguards 3 (v8)")
    v8_ig1_col        = _opt_col(headers, "v8 IG1")
    v8_ig2_col        = _opt_col(headers, "v8 IG2")
    v8_ig3_col        = _opt_col(headers, "v8 IG3")
    sg_v7_ig1_col     = _opt_col(headers, "CIS Safeguards 1 (v7)")
    sg_v7_ig2_col     = _opt_col(headers, "CIS Safeguards 2 (v7)")
    sg_v7_ig3_col     = _opt_col(headers, "CIS Safeguards 3 (v7)")
    v7_ig1_col        = _opt_col(headers, "v7 IG1")
    v7_ig2_col        = _opt_col(headers, "v7 IG2")
    v7_ig3_col        = _opt_col(headers, "v7 IG3")
    references_col    = _opt_col(headers, "References")
    default_value_col = _opt_col(headers, "Default Value")

    # Section headers in the XLSX have a title row above each group of recs.
    # Track current section name (non-numeric Title rows with no Recommendation #).
    current_section = ""
    level_map: dict[str, str] = {}   # rec_id -> level

    # First pass: read level sheets to get L1/L2 and license per rec
    license_map: dict[str, str] = {}
    for sheet_name in wb.sheetnames:
        m = re.match(r'^(E\d+) Level (\d+)$', sheet_name)
        if not m:
            continue
        lic, lvl = m.group(1), m.group(2)
        ws2 = wb[sheet_name]
        sh_headers = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
        try:
            r_col = sh_headers.index("Recommendation #")
        except ValueError:
            continue
        for row2 in ws2.iter_rows(min_row=2, values_only=True):
            r = str(row2[r_col]).strip() if row2[r_col] else ""
            if re.match(r'^\d+\.\d+', r):
                if r not in level_map:
                    level_map[r] = f"L{lvl}"
                    license_map[r] = lic

    controls: dict = {}
    titles: dict = {}

    for row in rows[1:]:
        rec = str(row[rec_col]).strip() if row[rec_col] else ""
        raw_title = str(row[title_col]).strip() if row[title_col] else ""

        # Section header row (no rec number)
        if not re.match(r'^\d+\.\d+', rec):
            if raw_title and rec == "":
                current_section = raw_title
            continue

        title = re.sub(r'^\(L\d\)\s*', '', raw_title).strip()
        safeguards = [
            str(row[i]).strip()
            for i in sg_cols
            if row[i] and str(row[i]).strip() not in ("", "None")
        ]

        # Union NIST IDs across all mapped safeguards
        nist_seen: set[str] = set()
        nist_ids: list[str] = []
        for sg in safeguards:
            for nist_id in sg_nist.get(sg, []):
                if nist_id not in nist_seen:
                    nist_seen.add(nist_id)
                    nist_ids.append(nist_id)

        # Phase 1 enrichment fields (#347)
        section_number = None
        if section_num_col is not None and row[section_num_col] is not None:
            try:
                section_number = int(str(row[section_num_col]).strip())
            except ValueError:
                section_number = None
        # Fallback: derive from rec# leading digits
        if section_number is None:
            sm = re.match(r'^(\d+)', rec)
            section_number = int(sm.group(1)) if sm else None

        assessment_status = _cell_text(row[assess_col]) if assess_col is not None else ""
        if assessment_status not in ("Manual", "Automated"):
            assessment_status = ""

        cis_controls_v8 = _parse_cis_controls_cell(row[cis_controls_col]) if cis_controls_col is not None else []

        v8_ig1_sgs = _parse_safeguards_cell(row[sg_v8_ig1_col]) if sg_v8_ig1_col is not None else []
        v8_ig2_sgs = _parse_safeguards_cell(row[sg_v8_ig2_col]) if sg_v8_ig2_col is not None else []
        v8_ig3_sgs = _parse_safeguards_cell(row[sg_v8_ig3_col]) if sg_v8_ig3_col is not None else []
        v8_applicable: list[str] = []
        if v8_ig1_col is not None and _parse_ig_flag(row[v8_ig1_col]): v8_applicable.append("IG1")
        if v8_ig2_col is not None and _parse_ig_flag(row[v8_ig2_col]): v8_applicable.append("IG2")
        if v8_ig3_col is not None and _parse_ig_flag(row[v8_ig3_col]): v8_applicable.append("IG3")

        v7_ig1_sgs = _parse_safeguards_cell(row[sg_v7_ig1_col]) if sg_v7_ig1_col is not None else []
        v7_ig2_sgs = _parse_safeguards_cell(row[sg_v7_ig2_col]) if sg_v7_ig2_col is not None else []
        v7_ig3_sgs = _parse_safeguards_cell(row[sg_v7_ig3_col]) if sg_v7_ig3_col is not None else []
        v7_applicable: list[str] = []
        if v7_ig1_col is not None and _parse_ig_flag(row[v7_ig1_col]): v7_applicable.append("IG1")
        if v7_ig2_col is not None and _parse_ig_flag(row[v7_ig2_col]): v7_applicable.append("IG2")
        if v7_ig3_col is not None and _parse_ig_flag(row[v7_ig3_col]): v7_applicable.append("IG3")

        references = _parse_references(row[references_col]) if references_col is not None else []
        default_value = _cell_text(row[default_value_col]) if default_value_col is not None else ""

        if rec not in controls:
            entry: dict = {
                "title":      title,
                "section":    current_section,
                "level":      level_map.get(rec, ""),
                "license":    license_map.get(rec, ""),
                "safeguards": safeguards,
                "nist800_53": nist_ids,
            }
            # Only emit phase-1 fields when source XLSX populated them, so that
            # legacy XLSX builds (without these columns) still produce the v1.1
            # shape and consumers see the deliberate "field absent" signal.
            if section_number is not None:
                entry["sectionNumber"] = section_number
            if assessment_status:
                entry["assessmentStatus"] = assessment_status
            if cis_controls_v8:
                entry["cisControls"] = cis_controls_v8
            v8_payload: dict = {}
            if v8_ig1_sgs: v8_payload["ig1"] = v8_ig1_sgs
            if v8_ig2_sgs: v8_payload["ig2"] = v8_ig2_sgs
            if v8_ig3_sgs: v8_payload["ig3"] = v8_ig3_sgs
            if v8_applicable: v8_payload["applicableIGs"] = v8_applicable
            v7_payload: dict = {}
            if v7_ig1_sgs: v7_payload["ig1"] = v7_ig1_sgs
            if v7_ig2_sgs: v7_payload["ig2"] = v7_ig2_sgs
            if v7_ig3_sgs: v7_payload["ig3"] = v7_ig3_sgs
            if v7_applicable: v7_payload["applicableIGs"] = v7_applicable
            sgbyver: dict = {}
            if v8_payload: sgbyver["v8"] = v8_payload
            if v7_payload: sgbyver["v7"] = v7_payload
            if sgbyver:
                entry["cisSafeguardsByVersion"] = sgbyver
            if default_value:
                entry["defaultValue"] = default_value
            if references:
                entry["references"] = references
            controls[rec] = entry
            titles[rec] = title

    return controls, titles


# ---------------------------------------------------------------------------
# SCuBA NIST mapping
# ---------------------------------------------------------------------------

def build_scuba_mapping(csv_path: Path) -> dict[str, list[str]]:
    """{base_policy_id: [nist_ids]} from CISA ScubaGear NIST mapping CSV.

    Keys strip the version suffix (e.g. 'MS.AAD.3.2') so lookup works
    regardless of whether the registry stores v1 or v2.
    """
    mapping: dict[str, list[str]] = {}
    with open(csv_path, encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            policy_id = row.get("scuba-control-id", "").strip()
            nist_raw  = row.get("nist-800-53-control-id", "").strip()
            if not policy_id or not nist_raw:
                continue
            base_id = re.sub(r'v\d+$', '', policy_id)
            nist_ids = parse_nist_ids(nist_raw, sep=",")
            if base_id not in mapping:
                mapping[base_id] = nist_ids
            else:
                # Merge if same base ID appears under multiple versions
                existing = set(mapping[base_id])
                mapping[base_id].extend(n for n in nist_ids if n not in existing)
    return mapping


# ---------------------------------------------------------------------------
# Writers
# ---------------------------------------------------------------------------

def write_crosswalk(controls: dict, out_path: Path) -> None:
    payload = OrderedDict([
        ("schemaVersion", SCHEMA_VERSION),
        ("source",        f"{XLSX_NAME} + {SG_CSV_NAME}"),
        ("derivation",    "CIS M365 recommendation -> CIS Controls v8 safeguard -> NIST 800-53 R5"),
        ("phase1Note",    "Schema v1.2.0 adds factual metadata (sectionNumber, "
                          "assessmentStatus, cisControls, cisSafeguardsByVersion, "
                          "defaultValue, references) per #347 phase 1. CIS-authored "
                          "prose deferred pending licensing resolution."),
        ("controls",      controls),
    ])
    out_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    covered = sum(1 for c in controls.values() if c["nist800_53"])
    enriched = sum(1 for c in controls.values() if any(
        k in c for k in ("assessmentStatus", "cisControls", "cisSafeguardsByVersion", "defaultValue", "references")
    ))
    print(f"  Written: {out_path.relative_to(REPO_ROOT)} "
          f"({len(controls)} controls, {covered} with NIST mappings, "
          f"{enriched} with #347 phase-1 enrichment)")


def write_scuba_mapping(mapping: dict, out_path: Path) -> None:
    payload = OrderedDict([
        ("schemaVersion", "1.0.0"),
        ("source",        SCUBA_CSV_NAME),
        ("baseline",      "FedRAMP High"),
        ("derivation",    "CISA ScubaGear authoritative SCuBA policy -> NIST 800-53 R5 mapping"),
        ("controls",      mapping),
    ])
    out_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    print(f"  Written: {out_path.relative_to(REPO_ROOT)} ({len(mapping)} SCuBA policies)")


def update_framework_titles(titles: dict, titles_path: Path) -> None:
    existing = json.loads(titles_path.read_text(encoding="utf-8"))
    existing["cis-m365-v6"] = titles
    ordered = OrderedDict(sorted(existing.items()))
    titles_path.write_text(
        json.dumps(ordered, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    print(f"  Updated: {titles_path.relative_to(REPO_ROOT)} "
          f"(cis-m365-v6: {len(titles)} titles)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--cis-dir", type=Path, default=DEFAULT_CIS_DIR,
        help="Directory containing CIS source files (default: %(default)s)",
    )
    args = parser.parse_args()

    xlsx_path  = args.cis_dir / XLSX_NAME
    sg_csv     = args.cis_dir / SG_CSV_NAME
    scuba_csv  = args.cis_dir / SCUBA_CSV_NAME

    for p in (xlsx_path, sg_csv, scuba_csv):
        if not p.exists():
            print(f"ERROR: file not found: {p}")
            raise SystemExit(1)

    print(f"Reading safeguard -> NIST mapping: {sg_csv.name}")
    sg_nist = load_safeguard_nist(sg_csv)
    print(f"  {len(sg_nist)} safeguards with NIST mappings")

    print(f"Reading CIS M365 benchmark: {xlsx_path.name}")
    controls, titles = build_crosswalk(xlsx_path, sg_nist)
    print(f"  {len(controls)} recommendations parsed")

    print(f"Reading SCuBA NIST mapping: {scuba_csv.name}")
    scuba_mapping = build_scuba_mapping(scuba_csv)
    print(f"  {len(scuba_mapping)} SCuBA policies parsed")

    write_crosswalk(controls, CROSSWALK_OUT)
    write_scuba_mapping(scuba_mapping, SCUBA_MAPPING_OUT)
    update_framework_titles(titles, TITLES_PATH)

    print("\nDone.")


if __name__ == "__main__":
    main()
